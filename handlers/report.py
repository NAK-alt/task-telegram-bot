"""
Executive & Staff Reporting Module.
Generates on-demand reports with Telegram inline summaries or downloadable Excel (.xlsx) files.
Supports Boss (Executive) and Officer/Staff roles with strict data isolation.
"""

import io
import datetime
import pytz
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, InputFile
from telegram.ext import ContextTypes

import database as db
from i18n import t
from config import DEFAULT_TIMEZONE
from handlers.private import format_dt


async def report_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """
    Handle /report command or '📊 របាយការណ៍ (ប្រធាន)' / '📊 របាយការណ៍ (មន្ត្រី)' button press.
    """
    user = update.effective_user
    chat = update.effective_chat
    target_msg = update.effective_message

    if not user or not target_msg:
        return

    is_private = (chat.type == "private") if chat else True
    lang = db.get_user_language(user.id) if is_private else db.get_chat_language(chat.id if chat else user.id)

    cancel_btn = InlineKeyboardButton(t("btn_cancel", lang), callback_data="cancel_wizard")

async def report_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """
    Handle /report command or '📊 របាយការណ៍ (ប្រធាន)' / '📊 របាយការណ៍ (មន្ត្រី)' button press.
    """
    user = update.effective_user
    chat = update.effective_chat
    target_msg = update.effective_message

    if not user or not target_msg:
        return

    is_private = (chat.type == "private") if chat else True
    lang = db.get_user_language(user.id) if is_private else db.get_chat_language(chat.id if chat else user.id)

    cancel_btn = InlineKeyboardButton(t("btn_cancel", lang), callback_data="cancel_wizard")

    if not is_private:
        # In Group Chat: Hide personal tasks report completely! Show only Staff Assignments Report options for Boss.
        if db.is_boss(user.id):
            rows = [
                [InlineKeyboardButton(t("btn_fmt_msg", lang), callback_data="rpt_fmt_msg_staff")],
                [InlineKeyboardButton(t("btn_fmt_excel", lang), callback_data="rpt_fmt_excel_staff")],
            ]
            if user.id == db.ADMIN_USER_ID:
                rows.append([InlineKeyboardButton("⚠️ 🗑️ លុបទិន្នន័យរបាយការណ៍ទាំងអស់", callback_data="rpt_clear_all_confirm")])
            rows.append([cancel_btn])
            fmt_kb = InlineKeyboardMarkup(rows)
            await target_msg.reply_text(t("prompt_report_fmt", lang), reply_markup=fmt_kb)
        else:
            fmt_kb = InlineKeyboardMarkup([
                [
                    InlineKeyboardButton(t("btn_fmt_msg", lang), callback_data="rpt_fmt_msg_self"),
                ],
                [
                    InlineKeyboardButton(t("btn_fmt_excel", lang), callback_data="rpt_fmt_excel_self"),
                ],
                [cancel_btn]
            ])
            await target_msg.reply_text(t("prompt_report_staff_fmt", lang), reply_markup=fmt_kb)
        return

    # In Private Chat: Boss can choose between Personal To-Dos and Staff Member Assignments
    if db.is_boss(user.id):
        rows = [
            [InlineKeyboardButton(t("btn_rpt_personal", lang), callback_data="rpt_type_personal")],
            [InlineKeyboardButton(t("btn_rpt_staff", lang), callback_data="rpt_type_staff")],
        ]
        if user.id == db.ADMIN_USER_ID:
            rows.append([InlineKeyboardButton("📋 គ្រប់គ្រង & កែប្រែ/លុប របាយការណ៍ទាំងអស់ (Admin)", callback_data="rpt_admin_manage")])
            rows.append([InlineKeyboardButton("⚠️ 🗑️ លុបទិន្នន័យរបាយការណ៍ទាំងអស់", callback_data="rpt_clear_all_confirm")])
        rows.append([cancel_btn])
        inline_kb = InlineKeyboardMarkup(rows)
        await target_msg.reply_text(t("prompt_report_type", lang), reply_markup=inline_kb)
    else:
        fmt_kb = InlineKeyboardMarkup([
            [
                InlineKeyboardButton(t("btn_fmt_msg", lang), callback_data="rpt_fmt_msg_self"),
            ],
            [
                InlineKeyboardButton(t("btn_fmt_excel", lang), callback_data="rpt_fmt_excel_self"),
            ],
            [cancel_btn]
        ])
        await target_msg.reply_text(t("prompt_report_staff_fmt", lang), reply_markup=fmt_kb)


async def report_admin_manage_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """
    Admin Master Report Panel (Telegram ID 1079885088 only).
    Lists every single task/report in system with Edit (✏️) and Delete (🗑️) buttons for each item.
    """
    query = update.callback_query
    user = update.effective_user
    if not query or not user:
        return

    lang = db.get_user_language(user.id)
    if user.id != db.ADMIN_USER_ID:
        await query.answer(t("unauthorized_admin", lang), show_alert=True)
        return

    await query.answer()
    data = query.data

    if data == "rpt_admin_manage":
        tasks = db.get_all_system_tasks()
        if not tasks:
            empty_msg = (
                "✨ មិនទាន់មានរបាយការណ៍/ភារកិច្ចណាមួយនៅក្នុងប្រព័ន្ធនៅឡើយទេ។"
                if lang == "km" else
                "✨ No system reports or task records found."
            )
            await query.edit_message_text(empty_msg)
            return

        header = (
            "📋 **បញ្ជីរបាយការណ៍ & ភារកិច្ចទាំងអស់ក្នុងប្រព័ន្ធ (Admin Master Panel)៖**\n"
            if lang == "km" else
            "📋 **All System Reports & Task Records (Admin Master Panel):**\n"
        )
        response_lines = [header.strip()]
        keyboard = []

        for idx, task in enumerate(tasks, start=1):
            status_str = "⏳ កំពុងដំណើរការ" if task.get("status") == "pending" else "✅ បានបញ្ចប់"
            assignee = f"@{task['assigned_to_username']}" if task.get("assigned_to_username") else "Self/General"
            dl_str = format_dt(task.get("deadline"), lang=lang)

            line = f"{idx}. {status_str} **{task['title']}** [#{task['task_id']}]\n   👤 អ្នកទទួល: {assignee} | ⏰ កំណត់: {dl_str}"
            response_lines.append(line)

            btn_edit = InlineKeyboardButton(f"✏️ កែ #{task['task_id']}", callback_data=f"edit_task_menu_{task['task_id']}")
            btn_del = InlineKeyboardButton(f"🗑️ លុប #{task['task_id']}", callback_data=f"del_task_{task['task_id']}")
            keyboard.append([btn_edit, btn_del])

        keyboard.append([InlineKeyboardButton("⚠️ 🗑️ លុបទិន្នន័យទាំងអស់ (Clear All)", callback_data="rpt_clear_all_confirm")])
        keyboard.append([InlineKeyboardButton(t("btn_cancel", lang), callback_data="cancel_wizard")])

        inline_markup = InlineKeyboardMarkup(keyboard)
        await query.edit_message_text("\n\n".join(response_lines), reply_markup=inline_markup)

    elif data.startswith("edit_task_menu_"):
        task_id = data.replace("edit_task_menu_", "")
        task = db.get_task_by_id(task_id)
        if not task:
            await query.edit_message_text(t("todo_not_found", lang))
            return

        dl_str = format_dt(task.get("deadline"), lang=lang)
        prompt_text = (
            f"✏️ **កែប្រែភារកិច្ច/របាយការណ៍ #{task_id}៖**\n\n"
            f"📌 **ចំណងជើង៖** {task['title']}\n"
            f"⏰ **កាលបរិច្ឆេទកំណត់៖** {dl_str}\n"
            f"📊 **ស្ថានភាព៖** {task.get('status', 'pending')}\n\n"
            f"សូមជ្រើសរើសផ្នែកដែលត្រូវកែប្រែ៖"
        )
        keyboard = [
            [InlineKeyboardButton("✏️ កែប្រែចំណងជើង (Edit Title)", callback_data=f"edit_title_{task_id}")],
            [InlineKeyboardButton("⏰ កែប្រែកាលបរិច្ឆេទ (Edit Deadline)", callback_data=f"edit_dl_{task_id}")],
            [InlineKeyboardButton("⬅️ ត្រឡប់ក្រោយ (Back)", callback_data="rpt_admin_manage")]
        ]
        await query.edit_message_text(prompt_text, reply_markup=InlineKeyboardMarkup(keyboard))

    elif data.startswith("edit_title_"):
        task_id = data.replace("edit_title_", "")
        context.user_data["editing_task_id"] = task_id
        prompt_msg = (
            f"✏️ **សូមផ្ញើសារចំណងជើងថ្មីសម្រាប់ភារកិច្ច #${task_id} មកកាន់ទីនេះ៖**"
            if lang == "km" else
            f"✏️ **Please send the new title for task #${task_id} here:**"
        )
        cancel_btn = InlineKeyboardButton(t("btn_cancel", lang), callback_data="cancel_wizard")
        await query.edit_message_text(prompt_msg, reply_markup=InlineKeyboardMarkup([[cancel_btn]]))

    elif data.startswith("edit_dl_"):
        task_id = data.replace("edit_dl_", "")
        context.user_data["editing_dl_task_id"] = task_id
        from calendar_picker import create_calendar
        cal_kb = create_calendar()
        prompt_msg = (
            f"⏰ **សូមជ្រើសរើសកាលបរិច្ឆេទកំណត់ថ្មីសម្រាប់ភារកិច្ច #${task_id}៖**"
            if lang == "km" else
            f"⏰ **Please select new deadline date for task #${task_id}:**"
        )
        await query.edit_message_text(prompt_msg, reply_markup=cal_kb)


async def clear_all_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """
    Admin Handler: Delete all task records from Firestore.
    Restricted strictly to Telegram Admin ID 1079885088.
    """
    user = update.effective_user
    target_msg = update.effective_message
    if not user or not target_msg:
        return

    lang = db.get_user_language(user.id)
    if user.id != db.ADMIN_USER_ID:
        await target_msg.reply_text(t("unauthorized_admin", lang))
        return

    confirm_kb = InlineKeyboardMarkup([
        [
            InlineKeyboardButton("⚠️ បាទ/ចាស លុបទិន្នន័យទាំងអស់", callback_data="rpt_clear_all_do"),
        ],
        [
            InlineKeyboardButton(t("btn_cancel", lang), callback_data="cancel_wizard"),
        ]
    ])

    confirm_prompt = (
        "⚠️ **តើអ្នកពិតជាចង់លុបទិន្នន័យភារកិច្ច និងរបាយការណ៍ទាំងអស់ចេញពីប្រព័ន្ធមែនទេ?**\n\n*(កត់សម្គាល់៖ ប្រតិបត្តិការនេះនឹងលុបភារកិច្ចទាំងអស់ទាំងក្នុងដំណើរការ និងបានបញ្ចប់)*"
        if lang == "km" else
        "⚠️ **Are you sure you want to delete ALL tasks and reports from the system?**\n\n*(Note: This will delete all pending and completed task records)*"
    )
    await target_msg.reply_text(confirm_prompt, reply_markup=confirm_kb)


async def report_clear_all_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handle confirmation and execution of clearing all task records (Admin ID 1079885088 only)."""
    query = update.callback_query
    user = update.effective_user
    if not query or not user:
        return

    lang = db.get_user_language(user.id)
    if user.id != db.ADMIN_USER_ID:
        await query.answer(t("unauthorized_admin", lang), show_alert=True)
        return

    await query.answer()
    data = query.data

    if data == "rpt_clear_all_confirm":
        confirm_kb = InlineKeyboardMarkup([
            [
                InlineKeyboardButton("⚠️ បាទ/ចាស លុបទិន្នន័យទាំងអស់", callback_data="rpt_clear_all_do"),
            ],
            [
                InlineKeyboardButton(t("btn_cancel", lang), callback_data="cancel_wizard"),
            ]
        ])
        confirm_prompt = (
            "⚠️ **តើអ្នកពិតជាចង់លុបទិន្នន័យភារកិច្ច និងរបាយការណ៍ទាំងអស់ចេញពីប្រព័ន្ធមែនទេ?**\n\n*(កត់សម្គាល់៖ ប្រតិបត្តិការនេះនឹងលុបភារកិច្ចទាំងអស់ទាំងក្នុងដំណើរការ និងបានបញ្ចប់)*"
            if lang == "km" else
            "⚠️ **Are you sure you want to delete ALL tasks and reports from the system?**\n\n*(Note: This will delete all pending and completed task records)*"
        )
        await query.edit_message_text(confirm_prompt, reply_markup=confirm_kb)

    elif data == "rpt_clear_all_do":
        count = db.delete_all_tasks(user.id)
        done_text = (
            f"🗑️ **ទិន្នន័យភារកិច្ច និងរបាយការណ៍ទាំងអស់ (ចំនួន {count} ភារកិច្ច) ត្រូវបានលុបចេញពីប្រព័ន្ធដោយជោគជ័យ!**"
            if lang == "km" else
            f"🗑️ **All task records and reports ({count} tasks) have been successfully deleted from the system!**"
        )
        await query.edit_message_text(done_text)


async def report_type_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Prompt user to choose output format (Telegram vs Excel) for selected report type."""
    query = update.callback_query
    user = update.effective_user
    if not query or not user:
        return

    await query.answer()
    data = query.data
    lang = db.get_user_language(user.id)
    cancel_btn = InlineKeyboardButton(t("btn_cancel", lang), callback_data="cancel_wizard")

    report_type = data.replace("rpt_type_", "")  # 'personal' or 'staff'

    fmt_kb = InlineKeyboardMarkup([
        [
            InlineKeyboardButton(t("btn_fmt_msg", lang), callback_data=f"rpt_fmt_msg_{report_type}"),
        ],
        [
            InlineKeyboardButton(t("btn_fmt_excel", lang), callback_data=f"rpt_fmt_excel_{report_type}"),
        ],
        [cancel_btn]
    ])

    await query.edit_message_text(t("prompt_report_fmt", lang), reply_markup=fmt_kb)


async def report_generate_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Generate and deliver Telegram text summary or Excel document."""
    query = update.callback_query
    user = update.effective_user
    chat = update.effective_chat
    if not query or not user:
        return

    await query.answer()
    data = query.data
    lang = db.get_user_language(user.id) if (chat and chat.type == "private") else db.get_chat_language(chat.id if chat else user.id)

    # Parse callback string: e.g., rpt_fmt_msg_personal, rpt_fmt_excel_staff, rpt_fmt_msg_self
    parts = data.replace("rpt_fmt_", "").split("_", 1)  # ['msg', 'personal'] or ['excel', 'self']
    fmt = parts[0]  # 'msg' or 'excel'
    report_type = parts[1]  # 'personal', 'staff', or 'self'

    if report_type == "personal":
        if not db.is_boss(user.id):
            await query.edit_message_text(t("unauthorized_boss_only", lang))
            return
        tasks = db.get_all_personal_tasks(user.id)
        title = t("report_personal_title", lang)
        filename_prefix = "Boss_Personal_Tasks"

    elif report_type == "staff":
        if not db.is_boss(user.id):
            await query.edit_message_text(t("unauthorized_boss_only", lang))
            return
        tasks = db.get_all_assigned_by_boss(user.id)
        title = t("report_staff_title", lang)
        filename_prefix = "Staff_Assignments"

    else:  # report_type == 'self'
        tasks = db.get_all_tasks_for_staff(user.id, username=user.username or "")
        title = t("report_staff_self_title", lang)
        filename_prefix = f"Staff_Report_{user.username or user.id}"

    total = len(tasks)
    completed_count = sum(1 for t in tasks if t.get("status") == "completed")
    pending_count = total - completed_count
    pct = round((completed_count / total * 100), 1) if total > 0 else 0.0

    if fmt == "msg":
        # Generate Markdown Summary Message
        lines = [
            title,
            t("report_stat_summary", lang, total, completed_count, str(pct), pending_count),
            "---------------------------------------"
        ]

        if not tasks:
            lines.append("មិនមានទិន្នន័យភារកិច្ចនោះទេ។ (No task records found)")
        else:
            for task in tasks:
                status_icon = "✅" if task.get("status") == "completed" else "⏳"
                cr_str = format_dt(task.get("created_at"))
                dl_str = format_dt(task.get("deadline"))
                cm_str = format_dt(task.get("completed_at"))

                if report_type == "staff":
                    target_staff = f"@{task['assigned_to_username']}" if task.get("assigned_to_username") else "Unassigned"
                    line = f"{status_icon} {target_staff} - {task['title']}\n  📅 បង្កើត: {cr_str} | ⏰ កំណត់: {dl_str}"
                elif report_type == "self":
                    scope_tag = "(Personal)" if task.get("scope") == "private" else "(Assigned)"
                    line = f"{status_icon} {scope_tag} {task['title']}\n  📅 បង្កើត: {cr_str} | ⏰ កំណត់: {dl_str}"
                else:
                    line = f"{status_icon} {task['title']}\n  📅 បង្កើត: {cr_str} | ⏰ កំណត់: {dl_str}"

                if task.get("status") == "completed":
                    line += f" | ✅ បញ្ចប់: {cm_str}"
                lines.append(line)

        await query.edit_message_text("\n".join(lines))

    elif fmt == "excel":
        # Generate Excel (.xlsx) Document
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Task Report"

        # Styling Definitions
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9")
        )

        if report_type == "personal":
            headers = ["Task ID", "Title / Description", "Scope", "Status", "Created At", "Deadline", "Completed At"]
        elif report_type == "staff":
            headers = ["Task ID", "Assigned Staff", "Title / Description", "Status", "Assigned By", "Created At", "Deadline", "Completed At"]
        else:  # self
            headers = ["Task ID", "Task Type", "Title / Description", "Status", "Assigned By", "Created At", "Deadline", "Completed At"]

        ws.append(headers)

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Fill Data Rows
        for task in tasks:
            cr_str = format_dt(task.get("created_at"))
            dl_str = format_dt(task.get("deadline"))
            cm_str = format_dt(task.get("completed_at"))
            status_str = "Completed" if task.get("status") == "completed" else "Pending"

            if report_type == "personal":
                row = [
                    task.get("task_id", ""),
                    task.get("title", ""),
                    task.get("scope", "private"),
                    status_str,
                    cr_str,
                    dl_str,
                    cm_str if status_str == "Completed" else "N/A"
                ]
            elif report_type == "staff":
                staff_name = f"@{task['assigned_to_username']}" if task.get("assigned_to_username") else "Unassigned"
                row = [
                    task.get("task_id", ""),
                    staff_name,
                    task.get("title", ""),
                    status_str,
                    task.get("assigned_by_username", "Boss"),
                    cr_str,
                    dl_str,
                    cm_str if status_str == "Completed" else "N/A"
                ]
            else:  # self
                type_tag = "Personal To-Do" if task.get("scope") == "private" else "Assigned by Boss"
                assigner = task.get("assigned_by_username", "Self") if task.get("scope") == "group" else "Self"
                row = [
                    task.get("task_id", ""),
                    type_tag,
                    task.get("title", ""),
                    status_str,
                    assigner,
                    cr_str,
                    dl_str,
                    cm_str if status_str == "Completed" else "N/A"
                ]

            ws.append(row)

        # Apply borders and auto-width
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center")

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        # Save workbook to memory buffer
        output_buffer = io.BytesIO()
        wb.save(output_buffer)
        output_buffer.seek(0)

        now_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{filename_prefix}_{now_str}.xlsx"

        await query.edit_message_text(f"✓ Excel document generated successfully. Sending {filename}...")

        await context.bot.send_document(
            chat_id=chat.id,
            document=InputFile(output_buffer, filename=filename),
            caption=f"📊 {title.strip()}\nTotal: {total} | Completed: {completed_count} ({pct}%) | Pending: {pending_count}"
        )
